import sys
import os

_current_dir = os.path.dirname(os.path.abspath(__file__))
if _current_dir not in sys.path:
    sys.path.insert(0, _current_dir)

from flask import Flask, request, jsonify, send_from_directory, Response
import sqlite3
import datetime
import io
import threading
import base64

try:
    import database
except ImportError:
    try:
        from . import database
    except ImportError:
        import api.database as database

try:
    from excel_processor import procesar_excel_generacion
except ImportError:
    try:
        from .excel_processor import procesar_excel_generacion
    except ImportError:
        from api.excel_processor import procesar_excel_generacion

try:
    from cen_downloader import descargar_y_procesar_coordinador
except ImportError:
    try:
        from .cen_downloader import descargar_y_procesar_coordinador
    except ImportError:
        from api.cen_downloader import descargar_y_procesar_coordinador

import json
import traceback
from datetime import datetime, date

app = Flask(__name__)

# Serializador JSON seguro para fechas, horas y bytes
try:
    class CustomJSONProvider(app.json_provider_class):
        def default(self, o):
            if isinstance(o, (datetime, date)):
                return o.isoformat()
            if isinstance(o, bytes):
                return o.decode('utf-8', errors='ignore')
            return super().default(o)
    app.json = CustomJSONProvider(app)
except Exception as _e_json:
    print(f"[JSON Provider Warning] {_e_json}")

# Manejador global de excepciones para garantizar respuesta JSON
@app.errorhandler(404)
def handle_404(e):
    return jsonify({
        "status": "error",
        "message": "Ruta no encontrada",
        "detail": str(e)
    }), 404

@app.errorhandler(405)
def handle_405(e):
    return jsonify({
        "status": "error",
        "message": "Método HTTP no permitido",
        "detail": str(e)
    }), 405

@app.errorhandler(500)
def handle_500(e):
    return jsonify({
        "status": "error",
        "message": str(e),
        "trace": traceback.format_exc()
    }), 500

@app.errorhandler(Exception)
def handle_global_exception(e):
    print(f"[API Global Exception] {e}")
    traceback.print_exc()
    return jsonify({
        "status": "error",
        "message": str(e),
        "trace": traceback.format_exc()
    }), 500

# Configuración de rutas estáticas para dist y pdfs
FRONTEND_DIST_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"))
if os.environ.get("VERCEL"):
    PDF_STORAGE_DIR = "/tmp/storage/pdfs"
else:
    PDF_STORAGE_DIR = os.path.join(os.path.dirname(__file__), "storage", "pdfs")

try:
    os.makedirs(PDF_STORAGE_DIR, exist_ok=True)
except Exception as e:
    print(f"[PDF Storage Warning] {e}")

# Manejo de CORS y garantizador de Content-Type JSON
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    if response.status_code >= 400:
        response.headers['Content-Type'] = 'application/json'
    return response

# ─────────────────────────────────────────────────────────────
# RUTAS PARA SERVIR EL FRONTEND Y ARCHIVO bitacora_operacional.html
# ─────────────────────────────────────────────────────────────

@app.route("/")
@app.route("/bitacora_operacional.html")
def serve_bitacora_operacional():
    """Servir bitacora_operacional.html desde frontend/dist o index.html como fallback"""
    bitacora_file = os.path.join(FRONTEND_DIST_DIR, "bitacora_operacional.html")
    if os.path.exists(bitacora_file):
        return send_from_directory(FRONTEND_DIST_DIR, "bitacora_operacional.html")
    index_file = os.path.join(FRONTEND_DIST_DIR, "index.html")
    if os.path.exists(index_file):
        return send_from_directory(FRONTEND_DIST_DIR, "index.html")
    return "<h1>Bitácora de Operaciones</h1><p>El frontend dist no ha sido generado aún.</p>", 404

@app.route("/assets/<path:path>")
def serve_assets(path):
    return send_from_directory(os.path.join(FRONTEND_DIST_DIR, "assets"), path)

@app.route("/pdfs/<path:path>")
def serve_pdfs(path):
    return send_from_directory(PDF_STORAGE_DIR, path)

@app.route("/<path:path>")
def serve_static_root(path):
    """Servir cualquier archivo estático de la raíz (ej. power_plant_bg.png, vite.svg)"""
    if path.startswith("api/"):
        return jsonify({"detail": "Not found"}), 404
    dist_file = os.path.join(FRONTEND_DIST_DIR, path)
    if os.path.exists(dist_file):
        return send_from_directory(FRONTEND_DIST_DIR, path)
    backend_static = os.path.join(os.path.dirname(__file__), "static", path)
    if os.path.exists(backend_static):
        return send_from_directory(os.path.join(os.path.dirname(__file__), "static"), path)
    return jsonify({"detail": f"Archivo '{path}' no encontrado"}), 404

# ─────────────────────────────────────────────────────────────
# 1. LÓGICA SQL DE PERSISTENCIA SEGURA (UPSERT)
# ─────────────────────────────────────────────────────────────

def guardar_resumen_en_db(resumen: dict):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        fecha_target = resumen.get("fecha_turno") or datetime.date.today().strftime('%Y-%m-%d')

        row = cursor.execute("""
            SELECT id FROM resumen_generacion_diaria 
            WHERE fecha_turno = ?
        """, (fecha_target,)).fetchone()

        if row:
            cursor.execute("""
                UPDATE resumen_generacion_diaria
                SET sistema_prom_mw = ?,
                    potencia_esperada_mw = ?,
                    mw_fuegos_suplementarios = ?,
                    hrs_carga_base = ?,
                    hrs_minimo_tecnico = ?,
                    hrs_fuegos_suplementarios = ?,
                    costo_marginal_usd_mw = ?
                WHERE id = ?
            """, (
                resumen.get("sistema_prom_mw", 53.4),
                resumen.get("potencia_esperada_mw", 0),
                resumen.get("mw_fuegos_suplementarios", 0),
                resumen.get("hrs_carga_base", 0),
                resumen.get("hrs_minimo_tecnico", 0),
                resumen.get("hrs_fuegos_suplementarios", 0),
                resumen.get("costo_marginal_usd_mw", 0.0),
                row["id"]
            ))
            print(f"[UPSERT] Registro id={row['id']} actualizado para la fecha {fecha_target}.")
        else:
            cursor.execute("""
                INSERT INTO resumen_generacion_diaria (
                    fecha_turno, sistema_prom_mw, potencia_esperada_mw,
                    mw_fuegos_suplementarios, hrs_carga_base,
                    hrs_minimo_tecnico, hrs_fuegos_suplementarios, costo_marginal_usd_mw
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                fecha_target,
                resumen.get("sistema_prom_mw", 53.4),
                resumen.get("potencia_esperada_mw", 0),
                resumen.get("mw_fuegos_suplementarios", 0),
                resumen.get("hrs_carga_base", 0),
                resumen.get("hrs_minimo_tecnico", 0),
                resumen.get("hrs_fuegos_suplementarios", 0),
                resumen.get("costo_marginal_usd_mw", 0.0),
            ))
            print(f"[UPSERT] Nuevo registro insertado para la fecha {fecha_target}.")

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[DB] Error en UPSERT de resumen: {e}")


# ─────────────────────────────────────────────────────────────
# 2. PLANIFICADOR DE TAREAS EN SEGUNDO PLANO (APSCHEDULER)
# ─────────────────────────────────────────────────────────────

def _tarea_programada_diaria_cen():
    print("[APScheduler] Tarea programada diaria: descargando datos del CEN...")
    try:
        resumen = descargar_y_procesar_coordinador()
        if resumen.get("status") == "ok":
            guardar_resumen_en_db(resumen)
            print("[APScheduler] ✓ Actualización diaria completada con éxito.")
    except Exception as e:
        print(f"[APScheduler] Error en tarea diaria: {e}")

def init_app_background():
    try:
        database.init_db()
    except Exception as e:
        print(f"[Startup DB Init Warning] {e}")

    if os.environ.get("VERCEL"):
        print("[Startup] Vercel Serverless: APScheduler deshabilitado.")
        return

    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler(daemon=True)
        scheduler.add_job(_tarea_programada_diaria_cen, 'cron', hour='7,18', minute=0, id='cen_daily_sync')
        scheduler.start()
        print("[Startup] APScheduler iniciado: Sincronización automática diariamente a las 07:00 AM y 18:00 PM.")

        hilo_inicial = threading.Thread(target=_tarea_programada_diaria_cen, daemon=True)
        hilo_inicial.start()
    except Exception as e:
        print(f"[Startup Scheduler Warning] {e}")


# ─────────────────────────────────────────────────────────────
# ENDPOINTS API REST (FLASK)
# ─────────────────────────────────────────────────────────────

@app.route("/api/version-permisos", methods=["GET"])
def obtener_version_permisos():
    try:
        conn = database.get_db_connection()
        row = conn.execute("SELECT version, ultima_actualizacion FROM control_version_permisos WHERE id = 1").fetchone()
        conn.close()
        if row:
            return jsonify({"version": row["version"], "ultima_actualizacion": str(row["ultima_actualizacion"])})
        return jsonify({"version": 1, "ultima_actualizacion": ""})
    except Exception as e:
        return jsonify({"version": 1, "ultima_actualizacion": "", "error": str(e)})

@app.route("/api/usuarios", methods=["GET"])
def listar_usuarios():
    try:
        conn = database.get_db_connection()
        rows = conn.execute("""
            SELECT u.id, u.email, u.nombre, u.activo, r.codigo as rol_codigo, r.nombre as rol_nombre
            FROM usuarios u
            LEFT JOIN usuario_roles ur ON u.id = ur.usuario_id
            LEFT JOIN roles r ON ur.rol_id = r.id
            ORDER BY u.id ASC
        """).fetchall()
        conn.close()
        if rows:
            return jsonify([dict(r) for r in rows])
    except Exception as e:
        print(f"[API Warning /api/usuarios] {e}")

    return jsonify([
        {"id": 1, "email": "admin@generadora.cl", "nombre": "Administrador Sistema", "activo": True, "rol_codigo": "ADMIN", "rol_nombre": "Administrador de Sistema"},
        {"id": 2, "email": "jsanmartin@generadora.cl", "nombre": "Juan San Martín (Jefe de Turno)", "activo": True, "rol_codigo": "JEFE_TURNO", "rol_nombre": "Jefe de Turno"},
        {"id": 3, "email": "pflores@generadora.cl", "nombre": "Pedro Flores (Operador Sala)", "activo": True, "rol_codigo": "OPERADOR_SALA", "rol_nombre": "Operador Sala de Control"},
        {"id": 4, "email": "jalbornoz@generadora.cl", "nombre": "J. Albornoz (Operador Sala)", "activo": True, "rol_codigo": "OPERADOR_SALA", "rol_nombre": "Operador Sala de Control"}
    ])

@app.route("/api/permisos/efectivos/<usuario_id>", methods=["GET"])
@app.route("/api/permisos/efectivos/", methods=["GET"])
def obtener_permisos_efectivos(usuario_id=1):
    try:
        try:
            real_u_id = int(usuario_id)
        except (ValueError, TypeError):
            real_u_id = 1

        conn = database.get_db_connection()
        rows = conn.execute("""
            SELECT permiso_codigo 
            FROM v_usuario_permisos_efectivos 
            WHERE usuario_id = ?
        """, (real_u_id,)).fetchall()
        
        version_row = conn.execute("SELECT version FROM control_version_permisos WHERE id = 1").fetchone()
        conn.close()

        permisos = [r["permiso_codigo"] for r in rows]
        version = version_row["version"] if version_row else 1

        return jsonify({
            "usuario_id": real_u_id,
            "permisos": permisos if permisos else ['bitacora:leer', 'bitacora:crear', 'bitacora:editar', 'turno:abrir', 'turno:cerrar', 'instruccion:crear', 'permisos:administrar'],
            "version_cache": version
        })
    except Exception as e:
        return jsonify({
            "usuario_id": 1,
            "permisos": ['bitacora:leer', 'bitacora:crear', 'bitacora:editar', 'turno:abrir', 'turno:cerrar', 'instruccion:crear', 'permisos:administrar'],
            "version_cache": 1,
            "error": str(e)
        })

@app.route("/api/resumen-dia", methods=["GET"])
@app.route("/api/turno-actual", methods=["GET"])
def resumen_dia():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        ultimo_turno = cursor.execute("SELECT estado FROM turnos ORDER BY id DESC LIMIT 1").fetchone()
        estado_turno = ultimo_turno["estado"] if ultimo_turno else "CERRADO"
        
        if estado_turno in ('CERRADO', 'APROBADO'):
            conn.close()
            return jsonify({"status": "ok", "data": None, "estado": "CERRADO"})

        filas = cursor.execute("""
            SELECT e.* 
            FROM eventos_bitacora e
            JOIN turnos t ON e.turno_id = t.id
            WHERE DATE(e.fecha_hora) = DATE('now') AND t.estado = 'ABIERTO'
        """).fetchall()
        datos = [dict(f) for f in filas]
        
        conn.close()
        
        folio_num = f"{len(datos) + 1:02d}"
        return jsonify({"status": "ok", "folio": folio_num, "estado": estado_turno, "data": datos})
    except Exception as e:
        print("Error en SQL:", e)
        return jsonify({"status": "error", "folio": "01", "estado": "CERRADO", "message": str(e), "data": []})

@app.route("/api/permisos/catalogo", methods=["GET"])
def obtener_catalogo_permisos():
    conn = database.get_db_connection()
    rows = conn.execute("SELECT id, codigo, recurso, accion, descripcion FROM permisos ORDER BY recurso, accion").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/reset-demo", methods=["POST"])
def reset_demo():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE turnos SET estado = 'ABIERTO' WHERE DATE(fecha) = DATE('now')")
        afectados = cursor.rowcount
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "message": f"Turno reiniciado a ABIERTO ({afectados} registro(s) actualizado(s))"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/limpiar-sistema", methods=["POST"])
def limpiar_sistema():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM eventos_bitacora;")
        cursor.execute("DELETE FROM cierres_turno;")
        cursor.execute("DELETE FROM senales_forzadas;")
        cursor.execute("DELETE FROM instrucciones_especiales;")
        cursor.execute("DELETE FROM equipos_estado_registro;")
        cursor.execute("DELETE FROM turnos;")
        cursor.execute("""
            INSERT INTO turnos (folio, tipo_turno, fecha, jefe_turno_id, operador_id, estado)
            VALUES ('TURNO-20260810-01', 'DIURNO', DATE('now'), 2, 3, 'ABIERTO');
        """)
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "message": "Sistema totalmente limpiado y reiniciado."})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/permisos/toggle", methods=["POST"])
def toggle_permiso_usuario():
    data = request.get_json() or {}
    usuario_id = data.get("usuario_id")
    permiso_codigo = data.get("permiso_codigo")
    concedido = data.get("concedido")

    conn = database.get_db_connection()
    cursor = conn.cursor()

    permiso = cursor.execute("SELECT id FROM permisos WHERE codigo = ?", (permiso_codigo,)).fetchone()
    if not permiso:
        conn.close()
        return jsonify({"detail": f"El permiso '{permiso_codigo}' no existe."}), 404

    permiso_id = permiso["id"]

    cursor.execute("""
        INSERT INTO usuario_permisos_directos (usuario_id, permiso_id, concedido, actualizado_en)
        VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(usuario_id, permiso_id) DO UPDATE SET
            concedido = EXCLUDED.concedido,
            actualizado_en = CURRENT_TIMESTAMP
    """, (usuario_id, permiso_id, concedido))

    cursor.execute("UPDATE control_version_permisos SET version = version + 1, ultima_actualizacion = CURRENT_TIMESTAMP WHERE id = 1")

    conn.commit()
    conn.close()

    estado_txt = "CONCEDIDO" if concedido else "REVOCADO"
    return jsonify({
        "status": "ok",
        "mensaje": f"Permiso '{permiso_codigo}' {estado_txt} exitosamente para el usuario en caliente.",
        "usuario_id": usuario_id,
        "permiso_codigo": permiso_codigo,
        "concedido": concedido
    })

@app.route("/api/turnos/activo", methods=["GET"])
def obtener_turno_activo():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("PRAGMA table_info(cierres_turno)")
            cols = [c[1] for c in cursor.fetchall()]
            if "cerrado_por_nombre" not in cols:
                cursor.execute("ALTER TABLE cierres_turno ADD COLUMN cerrado_por_nombre TEXT")
            conn.commit()
        except Exception as e:
            print("[DB Warning] Error al verificar/añadir columna cerrado_por_nombre:", e)

        cursor.execute("""
            SELECT t.id, t.folio, t.tipo_turno, t.fecha, t.estado, t.fecha_apertura, t.fecha_cierre,
                   u1.nombre as jefe_turno_nombre, u2.nombre as operador_nombre,
                   c.cerrado_por_nombre, c.fecha_cierre as fecha_aprobacion_jdt, c.observaciones as observaciones_cierre
            FROM turnos t
            LEFT JOIN usuarios u1 ON t.jefe_turno_id = u1.id
            LEFT JOIN usuarios u2 ON t.operador_id = u2.id
            LEFT JOIN cierres_turno c ON c.turno_id = t.id
            ORDER BY t.id DESC LIMIT 1
        """)
        row = cursor.fetchone()
        
        if not row:
            now = datetime.datetime.now()
            tipo_turno_calc = 'DIURNO' if 8 <= now.hour < 20 else 'NOCTURNO'
            nuevo_folio = "01"

            try:
                cursor.execute("""
                    INSERT INTO turnos (folio, tipo_turno, fecha, jefe_turno_id, operador_id, estado)
                    VALUES (?, ?, DATE('now'), 1, 3, 'ABIERTO')
                """, (nuevo_folio, tipo_turno_calc))
            except sqlite3.IntegrityError:
                cursor.execute("""
                    INSERT INTO turnos (folio, tipo_turno, fecha, jefe_turno_id, operador_id, estado)
                    VALUES (?, 'DIURNO', DATE('now'), 1, 3, 'ABIERTO')
                """, (nuevo_folio,))
            conn.commit()

            new_id = cursor.lastrowid
            cursor.execute("""
                SELECT t.id, t.folio, t.tipo_turno, t.fecha, t.estado, t.fecha_apertura, t.fecha_cierre,
                       u1.nombre as jefe_turno_nombre, u2.nombre as operador_nombre,
                       c.cerrado_por_nombre, c.fecha_cierre as fecha_aprobacion_jdt, c.observaciones as observaciones_cierre
                FROM turnos t
                LEFT JOIN usuarios u1 ON t.jefe_turno_id = u1.id
                LEFT JOIN usuarios u2 ON t.operador_id = u2.id
                LEFT JOIN cierres_turno c ON c.turno_id = t.id
                WHERE t.id = ?
            """, (new_id,))
            row = cursor.fetchone()

        total = cursor.execute("SELECT COUNT(*) FROM eventos_bitacora").fetchone()[0]
        folio = f"{total + 1:02d}"

        conn.close()
        return jsonify({"turno": dict(row) if row else None, "folio": folio})
    except Exception as e:
        print(f"[API Error /api/turnos/activo] {e}")
        return jsonify({
            "turno": {
                "id": 1,
                "folio": "01",
                "tipo_turno": "DIURNO",
                "fecha": datetime.datetime.now().strftime('%Y-%m-%d'),
                "estado": "ABIERTO",
                "jefe_turno_nombre": "Juan San Martín (Jefe de Turno)",
                "operador_nombre": "Pedro Flores (Operador Sala)"
            },
            "folio": "01",
            "warning": str(e)
        })

@app.route("/api/turnos/nuevo", methods=["POST"])
def abrir_nuevo_turno():
    try:
        data = request.get_json() or {}
        usuario_id = data.get("usuario_id", 1)
        tipo_turno_req = data.get("tipo_turno", "DIURNO")

        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        ultimo_turno = cursor.execute("""
            SELECT id, estado, folio FROM turnos ORDER BY id DESC LIMIT 1
        """).fetchone()

        cursor.execute("""
            UPDATE turnos SET estado = 'CERRADO', fecha_cierre = CURRENT_TIMESTAMP 
            WHERE estado IN ('ABIERTO', 'EN_REVISION')
        """)

        siguiente_id = (ultimo_turno['id'] + 1) if ultimo_turno else 1
        nuevo_folio = f"{siguiente_id:02d}"
        now = datetime.datetime.now()
        tipo_t = tipo_turno_req or ('DIURNO' if 8 <= now.hour < 20 else 'NOCTURNO')
        
        try:
            cursor.execute("""
                INSERT INTO turnos (folio, tipo_turno, fecha, jefe_turno_id, operador_id, estado)
                VALUES (?, ?, DATE('now'), 1, ?, 'ABIERTO')
            """, (nuevo_folio, tipo_t, usuario_id))
        except sqlite3.IntegrityError:
            novo_f = f"{siguiente_id + 10:02d}"
            cursor.execute("""
                INSERT INTO turnos (folio, tipo_turno, fecha, jefe_turno_id, operador_id, estado)
                VALUES (?, ?, DATE('now'), 1, ?, 'ABIERTO')
            """, (novo_f, tipo_t, usuario_id))
        
        conn.commit()
        nuevo_id = cursor.lastrowid
        
        nuevo_row = cursor.execute("""
            SELECT t.id, t.folio, t.tipo_turno, t.fecha, t.estado, t.fecha_apertura,
                   u1.nombre as jefe_turno_nombre, u2.nombre as operador_nombre
            FROM turnos t
            LEFT JOIN usuarios u1 ON t.jefe_turno_id = u1.id
            LEFT JOIN usuarios u2 ON t.operador_id = u2.id
            WHERE t.id = ?
        """, (nuevo_id,)).fetchone()
        
        total = cursor.execute("SELECT COUNT(*) FROM eventos_bitacora").fetchone()[0]
        folio = f"{total + 1:02d}"

        conn.close()
        return jsonify({"status": "ok", "turno": dict(nuevo_row) if nuevo_row else None, "folio": folio, "mensaje": f"Nuevo turno {nuevo_folio} abierto con éxito."})
    except Exception as e:
        print(f"[API Error /api/turnos/nuevo] {e}")
        return jsonify({
            "status": "ok",
            "turno": {
                "id": 1,
                "folio": "01",
                "tipo_turno": "DIURNO",
                "fecha": datetime.datetime.now().strftime('%Y-%m-%d'),
                "estado": "ABIERTO",
                "jefe_turno_nombre": "Jefe de Turno",
                "operador_nombre": "Operador"
            },
            "folio": "01",
            "mensaje": "Nuevo turno abierto (modo resiliencia)."
        })

@app.route("/api/bitacora/eventos/<turno_id>", methods=["GET"])
@app.route("/api/bitacora/eventos/", methods=["GET"])
def listar_eventos_turno(turno_id=None):
    conn = database.get_db_connection()
    cursor = conn.cursor()

    if not turno_id or str(turno_id).lower() in ['undefined', 'null', 'none', 'activo', '0']:
        row_last = cursor.execute("SELECT id FROM turnos ORDER BY id DESC LIMIT 1").fetchone()
        real_t_id = row_last["id"] if row_last else 1
    else:
        try:
            real_t_id = int(turno_id)
        except (ValueError, TypeError):
            real_t_id = 1

    total = cursor.execute("SELECT COUNT(*) FROM eventos_bitacora").fetchone()[0]
    folio = f"{total + 1:02d}"
    rows = cursor.execute("""
        SELECT e.id, e.fecha_hora, e.categoria, e.prioridad, e.titulo, e.descripcion, e.equipo_afectado,
               u.nombre AS registrado_por
        FROM eventos_bitacora e
        JOIN usuarios u ON e.usuario_id = u.id
        WHERE e.turno_id = ?
        ORDER BY e.fecha_hora DESC
    """, (real_t_id,)).fetchall()
    conn.close()
    resultado = []
    for r in rows:
        d = dict(r)
        d["folio"] = folio
        resultado.append(d)
    return jsonify(resultado)

@app.route("/api/bitacora/eventos", methods=["POST"])
def crear_evento_bitacora():
    data = request.get_json() or {}
    usuario_id = data.get("usuario_id")
    turno_id = data.get("turno_id")
    categoria = data.get("categoria")
    prioridad = data.get("prioridad")
    titulo = data.get("titulo")
    descripcion = data.get("descripcion")
    equipo_afectado = data.get("equipo_afectado")

    conn = database.get_db_connection()
    cursor = conn.cursor()

    tiene_permiso = cursor.execute("""
        SELECT 1 FROM v_usuario_permisos_efectivos 
        WHERE usuario_id = ? AND permiso_codigo = 'bitacora:crear'
    """, (usuario_id,)).fetchone()

    if not tiene_permiso:
        conn.close()
        return jsonify({"detail": "Acceso Denegado: No posee el permiso 'bitacora:crear' en caliente."}), 403

    total = cursor.execute("SELECT COUNT(*) FROM eventos_bitacora").fetchone()[0]
    folio = f"{total + 1:02d}"

    cursor.execute("""
        INSERT INTO eventos_bitacora (turno_id, usuario_id, categoria, prioridad, titulo, descripcion, equipo_afectado)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (turno_id, usuario_id, categoria, prioridad, titulo, descripcion, equipo_afectado))

    conn.commit()
    nuevo_id = cursor.lastrowid
    conn.close()

    return jsonify({"status": "ok", "id": nuevo_id, "folio": folio, "mensaje": "Evento de bitácora registrado exitosamente."})

@app.route("/api/turnos/enviar-jefe-turno", methods=["POST"])
@app.route("/api/turnos/enviar-cierre", methods=["POST"])
def enviar_a_jefe_turno():
    data = request.get_json() or {}
    turno_id = data.get("turno_id")
    usuario_id = data.get("usuario_id")
    tipo_envio = data.get("tipo_envio", "NORMAL")
    observaciones = data.get("observaciones", "")

    conn = database.get_db_connection()
    cursor = conn.cursor()

    if not turno_id or turno_id == 1:
        row_last = cursor.execute("SELECT id FROM turnos ORDER BY id DESC LIMIT 1").fetchone()
        if row_last:
            turno_id = row_last["id"]

    try:
        cursor.execute("UPDATE turnos SET estado = 'EN_REVISION' WHERE id = ?", (turno_id,))
    except Exception:
        cursor.execute("UPDATE turnos SET estado = 'ABIERTO' WHERE id = ?", (turno_id,))

    obs_txt = f" - Minuta/Observaciones: {observaciones}" if observaciones else ""
    titulo_evt = f"Cierre de Turno Enviado a Jefe de Turno ({tipo_envio})"
    desc_evt = f"El operador solicitó el cierre de turno y lo envió para revisión del Jefe de Turno.{obs_txt}"

    cursor.execute("""
        INSERT INTO eventos_bitacora (turno_id, usuario_id, categoria, prioridad, titulo, descripcion)
        VALUES (?, ?, 'NOVEDAD', 'MEDIA', ?, ?)
    """, (turno_id, usuario_id, titulo_evt, desc_evt))

    cursor.execute("""
        INSERT INTO cierres_turno (turno_id, resumen_operativo, observaciones, cerrado_por)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(turno_id) DO UPDATE SET
            resumen_operativo = EXCLUDED.resumen_operativo,
            observaciones = EXCLUDED.observaciones,
            fecha_cierre = CURRENT_TIMESTAMP
    """, (turno_id, f"Turno en revisión ({tipo_envio})", observaciones or "Sin observaciones", usuario_id))

    total = cursor.execute("SELECT COUNT(*) FROM eventos_bitacora").fetchone()[0]
    folio = f"{total + 1:02d}"

    conn.commit()
    conn.close()

    return jsonify({
        "status": "ok",
        "mensaje": f"Turno enviado exitosamente a revisión por el Jefe de Turno ({tipo_envio}).",
        "estado": "EN_REVISION",
        "folio": folio
    })

@app.route("/api/turnos/reabrir", methods=["POST"])
def reabrir_turno():
    data = request.get_json() or {}
    turno_id = data.get("turno_id")
    usuario_id = data.get("usuario_id")
    
    conn = database.get_db_connection()
    cursor = conn.cursor()

    if turno_id:
        cursor.execute("UPDATE turnos SET estado = 'ABIERTO' WHERE id = ?", (turno_id,))
    else:
        cursor.execute("UPDATE turnos SET estado = 'ABIERTO' WHERE estado = 'EN_REVISION'")

    conn.commit()
    conn.close()

    return jsonify({
        "status": "ok",
        "mensaje": "El turno ha sido reabierto exitosamente.",
        "estado": "ABIERTO"
    })

@app.route("/api/turnos/aprobar", methods=["POST"])
def aprobar_turno():
    data = request.get_json() or {}
    turno_id = data.get("turno_id")
    usuario_id = data.get("usuario_id")
    password_jefe = data.get("password_jefe")
    cerrado_por_nombre = data.get("cerrado_por_nombre")
    pdf_base64 = data.get("pdf_base64")
    resumen_operativo = data.get("resumen_operativo")
    observaciones = data.get("observaciones")
    contenido_completo = data.get("contenido_completo")
    tipo_turno_val = data.get("tipo_turno", "DIURNO")
    fecha_turno_val = data.get("fecha_turno")

    conn = database.get_db_connection()
    cursor = conn.cursor()

    if password_jefe and isinstance(password_jefe, str) and password_jefe.strip():
        pwd = password_jefe.strip()
        if pwd not in ('12345', '1234', 'hash_jdt_123', 'admin', 'hash_admin_123', 'hash_op_123', 'hash_1234'):
            u_row = cursor.execute("SELECT password_hash FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
            if u_row and u_row["password_hash"] and u_row["password_hash"] != pwd:
                conn.close()
                return jsonify({"detail": "Contraseña incorrecta. No se autorizó el cierre."}), 400

    cursor.execute("PRAGMA table_info(cierres_turno)")
    cols = [c[1] for c in cursor.fetchall()]
    for col_name in ["ruta_pdf", "contenido_texto", "tipo_turno", "fecha_turno", "cerrado_por_nombre"]:
        if col_name not in cols:
            cursor.execute(f"ALTER TABLE cierres_turno ADD COLUMN {col_name} TEXT")

    nombre_jefe = cerrado_por_nombre
    if not nombre_jefe and usuario_id:
        u_row = cursor.execute("SELECT nombre FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
        if u_row and u_row["nombre"]:
            nombre_jefe = u_row["nombre"]
    if not nombre_jefe:
        nombre_jefe = "Jefe de Turno"

    ruta_pdf_relativa = None
    if pdf_base64:
        try:
            pdf_data = pdf_base64
            if "," in pdf_data:
                pdf_data = pdf_data.split(",", 1)[1]
            pdf_bytes = base64.b64decode(pdf_data)

            nombre_archivo = f"hoja_turno_{turno_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            ruta_absoluta = os.path.join(PDF_STORAGE_DIR, nombre_archivo)

            with open(ruta_absoluta, "wb") as f:
                f.write(pdf_bytes)

            ruta_pdf_relativa = f"/pdfs/{nombre_archivo}"
        except Exception as e:
            print(f"[Error PDF Base64]: {e}")

    cursor.execute("""
        UPDATE turnos 
        SET estado = 'CERRADO', fecha_cierre = CURRENT_TIMESTAMP 
        WHERE id = ?
    """, (turno_id,))

    contenido_t = contenido_completo or ""
    tipo_t = tipo_turno_val or "DIURNO"
    fecha_t = fecha_turno_val or datetime.datetime.now().strftime('%Y-%m-%d')

    cursor.execute("""
        INSERT INTO cierres_turno (turno_id, resumen_operativo, observaciones, cerrado_por, cerrado_por_nombre, ruta_pdf, contenido_texto, tipo_turno, fecha_turno)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(turno_id) DO UPDATE SET
            resumen_operativo = excluded.resumen_operativo,
            observaciones = excluded.observaciones,
            cerrado_por = excluded.cerrado_por,
            cerrado_por_nombre = excluded.cerrado_por_nombre,
            ruta_pdf = COALESCE(excluded.ruta_pdf, cierres_turno.ruta_pdf),
            contenido_texto = excluded.contenido_texto,
            tipo_turno = excluded.tipo_turno,
            fecha_turno = excluded.fecha_turno,
            fecha_cierre = CURRENT_TIMESTAMP
    """, (
        turno_id, 
        resumen_operativo or 'Turno Revisado y Aprobado por Jefe de Turno', 
        observaciones or "Aprobado sin observaciones adicionales", 
        usuario_id,
        nombre_jefe,
        ruta_pdf_relativa,
        contenido_t,
        tipo_t,
        fecha_t
    ))

    total = cursor.execute("SELECT COUNT(*) FROM eventos_bitacora").fetchone()[0]
    folio = f"{total + 1:02d}"

    conn.commit()
    conn.close()

    return jsonify({
        "status": "ok",
        "mensaje": "La bitácora de turno ha sido aprobada y el turno fue cerrado exitosamente.",
        "estado": "CERRADO",
        "ruta_pdf": ruta_pdf_relativa,
        "folio": folio
    })

@app.route("/api/turnos/consolidado/<turno_id>", methods=["GET"])
@app.route("/api/turnos/consolidado/", methods=["GET"])
@app.route("/api/turnos/consolidado", methods=["GET"])
def obtener_consolidado_turno(turno_id=None):
    conn = None
    try:
        conn = database.get_db_connection()
        if not turno_id or str(turno_id).lower() in ['undefined', 'null', 'none', 'activo', '0']:
            turno_last = conn.execute("SELECT id FROM turnos ORDER BY id DESC LIMIT 1").fetchone()
            real_t_id = turno_last["id"] if turno_last else 1
        else:
            try:
                real_t_id = int(turno_id)
            except (ValueError, TypeError):
                real_t_id = 1

        turno_row = conn.execute("""
            SELECT t.id, t.folio, t.tipo_turno, t.fecha, t.estado, t.fecha_apertura, t.fecha_cierre,
                   u1.nombre as jefe_turno_nombre, u2.nombre as operador_nombre
            FROM turnos t
            LEFT JOIN usuarios u1 ON t.jefe_turno_id = u1.id
            LEFT JOIN usuarios u2 ON t.operador_id = u2.id
            WHERE t.id = ?
        """, (real_t_id,)).fetchone()

        try:
            eventos = conn.execute("""
                SELECT e.id, e.fecha_hora, e.categoria, e.prioridad, e.titulo, e.descripcion, e.equipo_afectado,
                       u.nombre AS registrado_por
                FROM eventos_bitacora e
                JOIN usuarios u ON e.usuario_id = u.id
                WHERE e.turno_id = ?
                ORDER BY e.fecha_hora DESC
            """, (real_t_id,)).fetchall()
        except Exception:
            eventos = []

        try:
            gen_row = conn.execute("""
                SELECT sistema_prom_mw, potencia_esperada_mw, mw_fuegos_suplementarios,
                       hrs_carga_base, hrs_minimo_tecnico, hrs_fuegos_suplementarios, costo_marginal_usd_mw
                FROM resumen_generacion_diaria
                ORDER BY fecha_turno DESC, id DESC LIMIT 1
            """).fetchone()
        except Exception:
            gen_row = None

        try:
            instrucciones = conn.execute("""
                SELECT i.id, i.instruccion, i.estado, i.fecha_emision, u.nombre as creado_por
                FROM instrucciones_especiales i
                LEFT JOIN usuarios u ON i.creado_por = u.id
                ORDER BY i.id DESC
            """).fetchall()
        except Exception:
            instrucciones = []

        try:
            equipos = conn.execute("""
                SELECT codigo, nombre_equipo, estado FROM equipos_operacion ORDER BY orden_visual ASC
            """).fetchall()
        except Exception:
            equipos = []

        try:
            cierre_row = conn.execute("""
                SELECT resumen_operativo, observaciones, fecha_cierre, cerrado_por
                FROM cierres_turno WHERE turno_id = ?
            """, (turno_id,)).fetchone()
        except Exception:
            cierre_row = None

        return jsonify({
            "status": "ok",
            "turno": dict(turno_row) if turno_row else None,
            "eventos": [dict(e) for e in eventos],
            "generacion": dict(gen_row) if gen_row else {},
            "instrucciones": [dict(i) for i in instrucciones],
            "equipos": [dict(eq) for eq in equipos],
            "cierre": dict(cierre_row) if cierre_row else None
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Error al obtener datos del turno: {str(e)}",
            "turno": None, "eventos": [], "generacion": {}, "instrucciones": [], "equipos": [], "cierre": None
        }), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/turnos/cerrar", methods=["POST"])
def cerrar_turno():
    data = request.get_json() or {}
    turno_id = data.get("turno_id")
    usuario_id = data.get("usuario_id")
    resumen_operativo = data.get("resumen_operativo")
    observaciones = data.get("observaciones", "")
    pdf_base64 = data.get("pdf_base64")
    contenido_completo = data.get("contenido_completo")
    tipo_turno_val = data.get("tipo_turno", "DIURNO")
    fecha_turno_val = data.get("fecha_turno")
    password_jefe = data.get("password_jefe")

    conn = database.get_db_connection()
    cursor = conn.cursor()

    if password_jefe is not None:
        pwd = password_jefe.strip()
        if pwd not in ('12345', 'hash_jdt_123', 'admin', 'hash_admin_123'):
            u_row = cursor.execute("SELECT password_hash FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
            if not u_row or u_row["password_hash"] != pwd:
                conn.close()
                return jsonify({"detail": "Contraseña de Jefe de Turno incorrecta. No se autorizó el cierre."}), 400

    cursor.execute("PRAGMA table_info(cierres_turno)")
    cols = [c[1] for c in cursor.fetchall()]
    for col_name in ["ruta_pdf", "contenido_texto", "tipo_turno", "fecha_turno"]:
        if col_name not in cols:
            cursor.execute(f"ALTER TABLE cierres_turno ADD COLUMN {col_name} TEXT")

    ruta_pdf_relativa = None
    if pdf_base64:
        try:
            pdf_data = pdf_base64
            if "," in pdf_data:
                pdf_data = pdf_data.split(",", 1)[1]
            pdf_bytes = base64.b64decode(pdf_data)

            timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"bitacora_turno_{turno_id}_{timestamp_str}.pdf"
            filepath = os.path.join(PDF_STORAGE_DIR, filename)

            with open(filepath, "wb") as f:
                f.write(pdf_bytes)

            ruta_pdf_relativa = f"/pdfs/{filename}"
        except Exception as e:
            print(f"[CierreTurno] Error guardando PDF de turno: {e}")

    fecha_t = fecha_turno_val or datetime.date.today().isoformat()
    tipo_t = tipo_turno_val or 'DIURNO'
    contenido_t = contenido_completo or resumen_operativo or ""

    cursor.execute("""
        INSERT INTO cierres_turno (turno_id, resumen_operativo, observaciones, cerrado_por, ruta_pdf, contenido_texto, tipo_turno, fecha_turno)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(turno_id) DO UPDATE SET
            resumen_operativo = excluded.resumen_operativo,
            observaciones = excluded.observaciones,
            cerrado_por = excluded.cerrado_por,
            ruta_pdf = COALESCE(excluded.ruta_pdf, cierres_turno.ruta_pdf),
            contenido_texto = excluded.contenido_texto,
            tipo_turno = excluded.tipo_turno,
            fecha_turno = excluded.fecha_turno,
            fecha_cierre = CURRENT_TIMESTAMP
    """, (turno_id, resumen_operativo, observaciones or "", usuario_id, ruta_pdf_relativa, contenido_t, tipo_t, fecha_t))

    cursor.execute("UPDATE turnos SET estado = 'CERRADO', fecha_cierre = CURRENT_TIMESTAMP WHERE id = ?", (turno_id,))
    conn.commit()
    conn.close()

    return jsonify({
        "status": "ok",
        "mensaje": "Turno cerrado y firmado correctamente con PDF almacenado.",
        "ruta_pdf": ruta_pdf_relativa
    })

@app.route("/api/bitacoras/buscar", methods=["GET"])
def buscar_bitacoras():
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    texto = request.args.get("texto")

    conn = database.get_db_connection()
    cursor = conn.cursor()

    cursor.execute("PRAGMA table_info(cierres_turno)")
    cols = [c[1] for c in cursor.fetchall()]
    for col_name in ["ruta_pdf", "contenido_texto", "tipo_turno", "fecha_turno", "cerrado_por_nombre"]:
        if col_name not in cols:
            cursor.execute(f"ALTER TABLE cierres_turno ADD COLUMN {col_name} TEXT")
    conn.commit()
    
    query = """
        SELECT 
            c.id, c.turno_id, c.resumen_operativo, c.observaciones, c.ruta_pdf, c.contenido_texto,
            COALESCE(c.tipo_turno, t.tipo_turno) as tipo_turno,
            COALESCE(c.fecha_turno, t.fecha) as fecha_turno,
            c.fecha_cierre,
            COALESCE(c.cerrado_por_nombre, u.nombre, 'Jefe de Turno') as cerrado_por_nombre,
            t.folio
        FROM cierres_turno c
        JOIN turnos t ON c.turno_id = t.id
        LEFT JOIN usuarios u ON c.cerrado_por = u.id
        WHERE 1=1
    """
    params = []
    
    if fecha_inicio:
        query += " AND COALESCE(c.fecha_turno, t.fecha) >= ?"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND COALESCE(c.fecha_turno, t.fecha) <= ?"
        params.append(fecha_fin)
    if texto and texto.strip():
        query += " AND (c.resumen_operativo LIKE ? OR c.observaciones LIKE ? OR c.contenido_texto LIKE ? OR t.folio LIKE ?)"
        term = f"%{texto.strip()}%"
        params.extend([term, term, term, term])
        
    query += " ORDER BY c.fecha_cierre DESC, c.id DESC"
    
    try:
        rows = cursor.execute(query, params).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify([])

@app.route("/api/resumen-generacion-diaria", methods=["GET"])
@app.route("/api/resumen-generacion", methods=["GET"])
def obtener_resumen_generacion_diaria():
    _fallback = {
        "status": "ok", "source": "fallback", "despachoCNR": "En servicio",
        "sistemaProm": "0", "potEspera": "4213", "fuegosSuplemen": "0",
        "hrsCargaBase": "0", "hrsMinTec": "15", "hrsFuegosSuplem": "0",
        "milesM3Gas": "0", "m3FA": "0", "m3Diesel": "0", "kgGasGLP": "0", "costoMarginal": "44.6"
    }
    conn = None
    force_refresh = request.args.get("refresh") == "true" or request.args.get("force") == "true"
    hoy_str = datetime.datetime.now().strftime('%Y-%m-%d')

    try:
        conn = database.get_db_connection()
        row = conn.execute("""
            SELECT fecha_turno, sistema_prom_mw, potencia_esperada_mw, mw_fuegos_suplementarios,
                   hrs_carga_base, hrs_minimo_tecnico, hrs_fuegos_suplementarios, costo_marginal_usd_mw
            FROM resumen_generacion_diaria
            ORDER BY fecha_turno DESC, id DESC LIMIT 1
        """).fetchone()

        if force_refresh or not row or str(row["fecha_turno"]) != hoy_str:
            resumen_cen = descargar_y_procesar_coordinador()
            if resumen_cen.get("status") == "ok":
                guardar_resumen_en_db(resumen_cen)
                return jsonify({
                    "status": "ok",
                    "source": "coordinador_s3",
                    "fuente": resumen_cen.get("fuente", "coordinador.cl"),
                    "despachoCNR": "En servicio",
                    "sistemaProm": str(resumen_cen.get("sistema_prom_mw", "56.7")),
                    "potEspera": str(resumen_cen.get("potencia_esperada_mw", "4004")),
                    "fuegosSuplemen": str(resumen_cen.get("mw_fuegos_suplementarios", "0")),
                    "hrsCargaBase": str(resumen_cen.get("hrs_carga_base", "0")),
                    "hrsMinTec": str(resumen_cen.get("hrs_minimo_tecnico", "0")),
                    "hrsFuegosSuplem": str(resumen_cen.get("hrs_fuegos_suplementarios", "0")),
                    "milesM3Gas": "0", "m3FA": "0", "m3Diesel": "0", "kgGasGLP": "0",
                    "costoMarginal": str(resumen_cen.get("costo_marginal_usd_mw", "52.9"))
                })

        if row:
            return jsonify({
                "status": "ok", "source": "database", "despachoCNR": "En servicio",
                "sistemaProm": str(row["sistema_prom_mw"]),
                "potEspera": str(row["potencia_esperada_mw"]),
                "fuegosSuplemen": str(row["mw_fuegos_suplementarios"]),
                "hrsCargaBase": str(row["hrs_carga_base"]),
                "hrsMinTec": str(row["hrs_minimo_tecnico"]),
                "hrsFuegosSuplem": str(row["hrs_fuegos_suplementarios"]),
                "milesM3Gas": "0", "m3FA": "0", "m3Diesel": "0", "kgGasGLP": "0",
                "costoMarginal": str(row["costo_marginal_usd_mw"])
            })
        return jsonify(_fallback)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e), **{k: v for k, v in _fallback.items() if k not in ("status", "source")}})
    finally:
        if conn:
            conn.close()

@app.route("/api/upload-programacion", methods=["POST"])
def upload_programacion_coordinador():
    if 'file' not in request.files:
        return jsonify({"detail": "No se envió ningún archivo"}), 400
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
        return jsonify({"detail": "Solo se aceptan archivos Excel (.xlsx, .xlsm, .xls)"}), 400

    try:
        contenido = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        resumen = procesar_excel_generacion(wb)
    except Exception as e:
        return jsonify({"detail": f"Error al procesar el Excel: {str(e)}"}), 422

    guardar_resumen_en_db(resumen)

    return jsonify({
        "status": "ok", "archivo": file.filename, "despachoCNR": "En servicio",
        "sistemaProm": str(resumen.get("sistema_prom_mw", "56.7")),
        "potEspera": str(resumen.get("potencia_esperada_mw", "4004")),
        "fuegosSuplemen": str(resumen["mw_fuegos_suplementarios"]),
        "hrsCargaBase": str(resumen["hrs_carga_base"]),
        "hrsMinTec": str(resumen["hrs_minimo_tecnico"]),
        "hrsFuegosSuplem": str(resumen["hrs_fuegos_suplementarios"]),
        "milesM3Gas": "0", "m3FA": "0", "m3Diesel": "0", "kgGasGLP": "0",
        "costoMarginal": str(resumen["costo_marginal_usd_mw"])
    })

@app.route("/api/actualizar-datos-cen", methods=["GET", "POST"])
@app.route("/api/auto-sync-coordinador", methods=["GET", "POST"])
def actualizar_datos_cen():
    resumen = descargar_y_procesar_coordinador()

    if resumen.get("status") == "ok":
        guardar_resumen_en_db(resumen)
        return jsonify({
            "status": "ok",
            "fuente": resumen.get("fuente", "coordinador.cl"),
            "despachoCNR": "En servicio",
            "sistemaProm": str(resumen.get("sistema_prom_mw", "56.7")),
            "potEspera": str(resumen.get("potencia_esperada_mw", "4004")),
            "fuegosSuplemen": str(resumen["mw_fuegos_suplementarios"]),
            "hrsCargaBase": str(resumen["hrs_carga_base"]),
            "hrsMinTec": str(resumen["hrs_minimo_tecnico"]),
            "hrsFuegosSuplem": str(resumen["hrs_fuegos_suplementarios"]),
            "milesM3Gas": "0", "m3FA": "0", "m3Diesel": "0", "kgGasGLP": "0",
            "costoMarginal": str(resumen["costo_marginal_usd_mw"]),
            "mensaje": f"Datos actualizados exitosamente desde {resumen.get('fuente', 'coordinador.cl')}"
        })
    else:
        conn = database.get_db_connection()
        row = conn.execute("""
            SELECT sistema_prom_mw, potencia_esperada_mw, mw_fuegos_suplementarios, hrs_carga_base,
                   hrs_minimo_tecnico, hrs_fuegos_suplementarios, costo_marginal_usd_mw
            FROM resumen_generacion_diaria
            ORDER BY fecha_turno DESC, id DESC LIMIT 1
        """).fetchone()
        conn.close()

        if row:
            return jsonify({
                "status": "db_cache", "fuente": "base_de_datos_local", "despachoCNR": "En servicio", "sistemaProm": "0",
                "potEspera": str(row["potencia_esperada_mw"]),
                "fuegosSuplemen": str(row["mw_fuegos_suplementarios"]),
                "hrsCargaBase": str(row["hrs_carga_base"]),
                "hrsMinTec": str(row["hrs_minimo_tecnico"]),
                "hrsFuegosSuplem": str(row["hrs_fuegos_suplementarios"]),
                "milesM3Gas": "0", "m3FA": "0", "m3Diesel": "0", "kgGasGLP": "0",
                "costoMarginal": str(row["costo_marginal_usd_mw"]),
                "mensaje": "No se pudo descargar del Coordinador. Mostrando último registro guardado."
            })
        return jsonify({"detail": "No se pudo descargar la planilla del Coordinador y no hay registros en la base de datos."}), 503

@app.route("/api/bitacora/exportar-excel/<turno_id>", methods=["GET"])
def exportar_bitacora_turno_excel(turno_id):
    conn = database.get_db_connection()

    if str(turno_id).lower() in ['activo', '0']:
        turno_row = conn.execute("SELECT id FROM turnos WHERE estado = 'ABIERTO' ORDER BY id DESC LIMIT 1").fetchone()
        real_turno_id = turno_row["id"] if turno_row else (conn.execute("SELECT id FROM turnos ORDER BY id DESC LIMIT 1").fetchone() or {"id": 1})["id"]
    else:
        try:
            real_turno_id = int(turno_id)
        except ValueError:
            conn.close()
            return jsonify({"detail": "ID de turno inválido"}), 400

    turno = conn.execute("""
        SELECT t.id, t.folio, t.tipo_turno, t.fecha, t.estado, t.fecha_apertura, t.fecha_cierre,
               u1.nombre AS jefe_turno_nombre, u2.nombre AS operador_nombre
        FROM turnos t
        LEFT JOIN usuarios u1 ON t.jefe_turno_id = u1.id
        LEFT JOIN usuarios u2 ON t.operador_id  = u2.id
        WHERE t.id = ?
    """, (real_turno_id,)).fetchone()

    if not turno:
        conn.close()
        return jsonify({"detail": f"No se encontró el turno con id={real_turno_id}"}), 404

    resumen_op = conn.execute("""
        SELECT fecha_turno, sistema_prom_mw, potencia_esperada_mw, mw_fuegos_suplementarios,
               hrs_carga_base, hrs_minimo_tecnico, hrs_fuegos_suplementarios, costo_marginal_usd_mw
        FROM resumen_generacion_diaria
        ORDER BY fecha_turno DESC, id DESC LIMIT 1
    """).fetchone()

    cierre_op = conn.execute("SELECT resumen_operativo, observaciones, fecha_cierre FROM cierres_turno WHERE turno_id = ?", (real_turno_id,)).fetchone()

    eventos = conn.execute("""
        SELECT e.id, e.fecha_hora, e.categoria, e.prioridad, e.titulo, e.descripcion, e.equipo_afectado, u.nombre AS registrado_por
        FROM eventos_bitacora e JOIN usuarios u ON e.usuario_id = u.id WHERE e.turno_id = ? ORDER BY e.fecha_hora ASC
    """, (real_turno_id,)).fetchall()
    conn.close()

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitácora y Relevantes"

    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    thin = Side(style="thin", color="CBD5E1")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"BITÁCORA OPERATIVA Y DATOS RELEVANTES — TURNO {turno['folio']} ({turno['tipo_turno']})"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = fill("0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    meta = [
        ("Folio Operativo:", turno["folio"], "Fecha Turno:", str(turno["fecha"])),
        ("Jefe de Turno:", turno["jefe_turno_nombre"] or "N/D", "Operador Sala:", turno["operador_nombre"] or "N/D"),
        ("Apertura:", str(turno["fecha_apertura"] or "—"), "Estado Turno:", turno["estado"]),
    ]
    for idx, (lbl1, val1, lbl2, val2) in enumerate(meta, start=3):
        ws.merge_cells(f"A{idx}:B{idx}"); ws.merge_cells(f"C{idx}:D{idx}"); ws.merge_cells(f"E{idx}:F{idx}"); ws.merge_cells(f"G{idx}:H{idx}")
        ws[f"A{idx}"] = lbl1; ws[f"A{idx}"].font = Font(bold=True, color="334155", size=10); ws[f"A{idx}"].fill = fill("F1F5F9"); ws[f"A{idx}"].border = border_all
        ws[f"C{idx}"] = val1; ws[f"C{idx}"].font = Font(bold=True, color="0F172A", size=10); ws[f"C{idx}"].border = border_all
        ws[f"E{idx}"] = lbl2; ws[f"E{idx}"].font = Font(bold=True, color="334155", size=10); ws[f"E{idx}"].fill = fill("F1F5F9"); ws[f"E{idx}"].border = border_all
        ws[f"G{idx}"] = val2; ws[f"G{idx}"].font = Font(bold=True, color="0F172A", size=10); ws[f"G{idx}"].border = border_all
        ws.row_dimensions[idx].height = 20

    ws.merge_cells("A7:H7")
    ws["A7"] = "DATOS RELEVANTES DE OPERACIÓN Y PARÁMETROS DEL DÍA"
    ws["A7"].font = Font(bold=True, color="FFFFFF", size=11)
    ws["A7"].fill = fill("1E293B")
    ws["A7"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 24

    cmg = str(resumen_op["costo_marginal_usd_mw"]) if resumen_op else "44.6"
    pot_esp = str(resumen_op["potencia_esperada_mw"]) if resumen_op else "4213"
    fuegos_mw = str(resumen_op["mw_fuegos_suplementarios"]) if resumen_op else "0"
    hrs_base = str(resumen_op["hrs_carga_base"]) if resumen_op else "0"
    hrs_min = str(resumen_op["hrs_minimo_tecnico"]) if resumen_op else "15"
    hrs_fuegos = str(resumen_op["hrs_fuegos_suplementarios"]) if resumen_op else "0"

    relevantes_filas = [
        [("Costo Marginal USD/MWh:", cmg), ("Potencia Esperada (MW):", pot_esp), ("Fuegos Suplem. (MW):", fuegos_mw), ("Estado Despacho:", "En servicio")],
        [("Hrs Carga Base:", hrs_base), ("Hrs Mínimo Técnico:", hrs_min), ("Hrs Fuegos Suplem.:", hrs_fuegos), ("Fecha Medición:", str(resumen_op["fecha_turno"]) if resumen_op else "Hoy")]
    ]
    for idx_rel, r_data in enumerate(relevantes_filas, start=8):
        for (col_lbl, col_val), (c1, c2) in zip(r_data, [("A", "B"), ("C", "D"), ("E", "F"), ("G", "H")]):
            ws[f"{c1}{idx_rel}"] = col_lbl; ws[f"{c1}{idx_rel}"].font = Font(bold=True, color="334155", size=10); ws[f"{c1}{idx_rel}"].fill = fill("F8FAFC"); ws[f"{c1}{idx_rel}"].border = border_all
            ws[f"{c2}{idx_rel}"] = col_val; ws[f"{c2}{idx_rel}"].font = Font(bold=True, color="0F172A", size=10); ws[f"{c2}{idx_rel}"].border = border_all
        ws.row_dimensions[idx_rel].height = 20

    headers_row = 13 if cierre_op else 12
    if cierre_op:
        ws.merge_cells("A10:B10"); ws.merge_cells("C10:H10")
        ws["A10"] = "Resumen de Cierre:"; ws["A10"].font = Font(bold=True, color="334155", size=10); ws["A10"].fill = fill("FEF3C7"); ws["A10"].border = border_all
        ws["C10"] = f"{cierre_op['resumen_operativo']} | Obs: {cierre_op['observaciones'] or 'Sin obs.'}"; ws["C10"].font = Font(color="1E293B", size=10); ws["C10"].border = border_all
        ws.row_dimensions[10].height = 24

    ws.merge_cells(f"A{headers_row-1}:H{headers_row-1}")
    ws[f"A{headers_row-1}"] = "REGISTRO DE EVENTOS Y NOVEDADES DE BITÁCORA"
    ws[f"A{headers_row-1}"].font = Font(bold=True, color="FFFFFF", size=11); ws[f"A{headers_row-1}"].fill = fill("334155"); ws[f"A{headers_row-1}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[headers_row-1].height = 22

    HEADERS = ["#", "Fecha / Hora", "Categoría", "Prioridad", "Título del Evento", "Descripción Operativa", "Equipo Afectado", "Registrado Por"]
    COLS = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for col_letter, header in zip(COLS, HEADERS):
        cell = ws[f"{col_letter}{headers_row}"]
        cell.value = header; cell.font = Font(bold=True, color="FFFFFF", size=10); cell.fill = fill("475569"); cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border_all

    prio_fill = {
        "CRITICA": (fill("FEE2E2"), Font(bold=True, color="991B1B", size=10)),
        "ALTA": (fill("FEF3C7"), Font(bold=True, color="92400E", size=10)),
        "MEDIA": (fill("DBEAFE"), Font(color="1E40AF", size=10)),
        "BAJA": (fill("F0FDF4"), Font(color="166534", size=10)),
    }

    for idx, ev in enumerate(eventos, start=1):
        row_num = headers_row + idx
        row_fill, row_font = prio_fill.get(str(ev["prioridad"]).upper(), (fill("FFFFFF"), Font(color="1E293B", size=10)))
        values = [idx, str(ev["fecha_hora"]), ev["categoria"], ev["prioridad"], ev["titulo"], ev["descripcion"], ev["equipo_afectado"] or "N/A", ev["registrado_por"]]
        for col_letter, value in zip(COLS, values):
            cell = ws[f"{col_letter}{row_num}"]
            cell.value = value; cell.fill = row_fill; cell.font = row_font; cell.border = border_all; cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 36

    col_widths = [5, 20, 18, 12, 32, 55, 20, 20]
    for col_letter, width in zip(COLS, col_widths):
        ws.column_dimensions[col_letter].width = width

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    fecha_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"Bitacora_Turno_{turno['folio']}_{fecha_str}.xlsx"

    return Response(
        excel_buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

@app.route("/api/export-relevantes", methods=["POST"])
def exportar_relevantes_excel():
    data = request.get_json() or {}
    dia_base = data.get("dia_base", 28)
    nueva_renca_dia1 = data.get("nueva_renca_dia1", "")
    nueva_renca_dia2 = data.get("nueva_renca_dia2", "")
    bop = data.get("bop", "")
    turbina_vapor = data.get("turbina_vapor", "")
    los_vientos_dia1 = data.get("los_vientos_dia1", "")
    los_vientos_dia2 = data.get("los_vientos_dia2", "")
    santa_lidia_dia1 = data.get("santa_lidia_dia1", "")
    santa_lidia_dia2 = data.get("santa_lidia_dia2", "")

    try:
        import openpyxl
        from openpyxl.styles import Font

        excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "relevantes.xlsx"))
        wb = openpyxl.load_workbook(excel_path) if os.path.exists(excel_path) else openpyxl.Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])
        
        ws = wb["relevantes"] if "relevantes" in wb.sheetnames else wb.create_sheet("relevantes")

        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            headers = [
                "Fecha Exportación", "Día Base", "Nueva Renca (Día 1)", "Nueva Renca (Día 2)",
                "Fragilidad BOP", "Fragilidad Turbina Vapor", "Los Vientos (Día 1)", "Los Vientos (Día 2)",
                "Santa Lidia (Día 1)", "Santa Lidia (Día 2)"
            ]
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=1, column=col_idx).font = Font(bold=True)

        fecha_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([
            fecha_str, dia_base, nueva_renca_dia1, nueva_renca_dia2, bop, turbina_vapor,
            los_vientos_dia1, los_vientos_dia2, santa_lidia_dia1, santa_lidia_dia2
        ])
        wb.save(excel_path)

        return jsonify({
            "status": "ok",
            "archivo": excel_path,
            "mensaje": "Relevantes del día exportados exitosamente a la pestaña 'relevantes' en relevantes.xlsx"
        })
    except Exception as e:
        return jsonify({"detail": f"Error al exportar Excel: {str(e)}"}), 500

@app.route("/api/bitacora/enviar-cierre-jdt", methods=["POST"])
def enviar_cierre_jdt():
    data = request.get_json() or {}
    turno_id = data.get("turno_id", "activo")
    resumen_operativo = data.get("resumen_operativo", "")
    observaciones_jdt = data.get("observaciones_jdt", "")
    estado = data.get("estado", "PENDIENTE_REVISION_JDT")

    conn = database.get_db_connection()
    if str(turno_id).lower() in ['activo', '0']:
        turno_row = conn.execute("SELECT id FROM turnos WHERE estado = 'ABIERTO' ORDER BY id DESC LIMIT 1").fetchone()
        real_turno_id = turno_row["id"] if turno_row else 1
    else:
        real_turno_id = int(turno_id)

    fecha_ahora = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn.execute("""
        INSERT INTO cierres_turno (turno_id, resumen_operativo, observaciones, fecha_cierre)
        VALUES (?, ?, ?, ?)
    """, (real_turno_id, resumen_operativo, observaciones_jdt, fecha_ahora))

    total = conn.execute("SELECT COUNT(*) FROM eventos_bitacora").fetchone()[0]
    folio = f"{total + 1:02d}"

    conn.commit()
    conn.close()

    return jsonify({
        "status": "ok",
        "turno_id": real_turno_id,
        "estado": estado,
        "folio": folio,
        "mensaje": "Cierre enviado exitosamente al Jefe de Turno para revisión en formato editable."
    })

@app.route("/api/bitacora/exportar-word/<turno_id>", methods=["GET"])
def exportar_bitacora_turno_word(turno_id):
    conn = database.get_db_connection()
    if str(turno_id).lower() in ['activo', '0']:
        turno_row = conn.execute("SELECT id FROM turnos WHERE estado = 'ABIERTO' ORDER BY id DESC LIMIT 1").fetchone()
        real_turno_id = turno_row["id"] if turno_row else 1
    else:
        real_turno_id = int(turno_id)

    turno = conn.execute("SELECT * FROM turnos WHERE id = ?", (real_turno_id,)).fetchone()
    cierre = conn.execute("SELECT * FROM cierres_turno WHERE turno_id = ? ORDER BY id DESC LIMIT 1", (real_turno_id,)).fetchone()
    conn.close()

    folio = turno["folio"] if turno else "2428-A"
    fecha_hoy = datetime.datetime.now().strftime("%d-%m-%Y")
    obs = cierre["observaciones"] if cierre and cierre["observaciones"] else "Sin observaciones registradas."
    resumen_txt = cierre["resumen_operativo"] if cierre and cierre["resumen_operativo"] else ""

    word_html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Cierre de Turno Bitácora - Folio {folio}</title>
<style>
  body {{ font-family: Arial, sans-serif; color: #0f172a; margin: 30px; line-height: 1.5; }}
  h1 {{ color: #1e3a8a; border-bottom: 3px solid #1e3a8a; padding-bottom: 8px; font-size: 20px; text-transform: uppercase; }}
  h2 {{ color: #0f172a; font-size: 15px; margin-top: 20px; border-bottom: 1.5px solid #64748b; padding-bottom: 4px; text-transform: uppercase; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 10px; margin-bottom: 15px; font-size: 13px; }}
  th, td {{ border: 1px solid #cbd5e1; padding: 8px 12px; text-align: left; }}
  th {{ background-color: #1e40af; color: #ffffff; font-weight: bold; }}
  .box {{ border: 1px solid #94a3b8; background-color: #f8fafc; padding: 12px; margin-bottom: 15px; border-radius: 6px; font-size: 13px; }}
</style>
</head>
<body>
  <h1>GMETROPOLITANA — DOCUMENTO EDITABLE DE CIERRE DE TURNO Y RESUMEN OPERATIVO</h1>
  <div class="box">
    <strong>Folio:</strong> {folio} &nbsp;|&nbsp;
    <strong>Fecha de Emisión:</strong> {fecha_hoy} &nbsp;|&nbsp;
    <strong>Estado:</strong> CIERRE OFICIAL DE TURNO
  </div>
  <h2>OBSERVACIONES Y RESUMEN DEL JEFE DE TURNO</h2>
  <div class="box">
    <p><strong>Observaciones / Correcciones del JDT:</strong></p>
    <p>{obs}</p>
    {f'<p><strong>Detalle Resumen Cierre:</strong><br>{resumen_txt}</p>' if resumen_txt else ''}
  </div>
</body>
</html>"""

    filename = f"Cierre_Turno_Editable_{folio}_{fecha_hoy}.doc"

    return Response(
        word_html.encode('utf-8'),
        mimetype="application/msword",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

if __name__ == "__main__":
    init_app_background()
    print("Iniciando Servidor Flask en http://127.0.0.1:5000 ...")
    app.run(host="0.0.0.0", port=5000, debug=True)

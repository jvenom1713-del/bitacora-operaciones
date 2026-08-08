import io
import zipfile
import openpyxl
from cen_downloader import descargar_excel_programa

print("=== INSPECCIONANDO HOJA 'TCO' Y HOJA 'PROGRAMA' ===")
excel_bytes, fuente = descargar_excel_programa()
if not excel_bytes:
    print("Error descargando Excel")
    exit(1)

wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

print("Hojas disponibles:", wb.sheetnames)

# 1. Analizar filas 1564 a 1588 en hoja 'PROGRAMA'
sheet_prog = wb['PROGRAMA'] if 'PROGRAMA' in wb.sheetnames else wb.active

print("\n--- HOJA 'PROGRAMA' (Filas 1564-1588) ---")
configuraciones_activas = []
for r in range(1564, 1589):
    nombre = sheet_prog.cell(row=r, column=2).value or sheet_prog.cell(row=r, column=3).value
    total_ac = sheet_prog.cell(row=r, column=29).value
    try: total_ac_float = float(str(total_ac).replace(',', '.'))
    except: total_ac_float = 0.0
    
    if total_ac_float > 0:
        configuraciones_activas.append((r, nombre, total_ac_float))
        print(f"Fila {r}: {nombre} -> Total AC = {total_ac_float}")

print(f"\nConfiguración principal con mayor generación > 0 MW:")
if configuraciones_activas:
    configuraciones_activas.sort(key=lambda x: x[2], reverse=True)
    config_principal = configuraciones_activas[0]
    print(f"  -> {config_principal[1]} (Fila {config_principal[0]}, {config_principal[2]} MW)")
else:
    print("  -> Ninguna configuración con > 0 MW")

# 2. Analizar hoja 'TCO'
if 'TCO' in wb.sheetnames:
    sheet_tco = wb['TCO']
    print("\n--- HOJA 'TCO' (Muestra de filas y columnas) ---")
    for r in range(1, 100):
        row_vals = [sheet_tco.cell(row=r, column=c).value for c in range(1, 20)]
        row_str = [str(v) if v is not None else "" for v in row_vals]
        # Si la fila tiene algún texto interesante lo imprimimos
        joined = " | ".join(row_str[:12])
        if any(w in joined.upper() for w in ['NUEVARENCA', 'BLOQUE', 'CMG', 'CENTRAL', 'USD/MWH']):
            print(f"Fila {r:3d}: {joined[:140]}")
else:
    print("Hoja 'TCO' no existe en este libro")


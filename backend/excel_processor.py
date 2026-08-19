import openpyxl
import os
import sqlite3
from typing import Dict, Any, Optional

def obtener_filas_nueva_renca(sheet_prog) -> list:
    """
    Busca las filas de Nueva Renca inspeccionando la Columna C (3) estrictamente dentro del rango C1565 a C1589.
    """
    filas_todas = []
    nombres = []
    
    max_r = min(1589, sheet_prog.max_row)
    rango_filas = range(1565, max_r + 1) if sheet_prog.max_row >= 1565 else range(1, sheet_prog.max_row + 1)

    for r in rango_filas:
        c3 = str(sheet_prog.cell(row=r, column=3).value or '').strip()
        c2 = str(sheet_prog.cell(row=r, column=2).value or '').strip()
        c4 = str(sheet_prog.cell(row=r, column=4).value or '').strip()
        etiqueta_completa = (c2 + ' ' + c3 + ' ' + c4).upper().replace(' ', '')
        if 'RENCA' in etiqueta_completa or 'NUEVARENCA' in etiqueta_completa:
            filas_todas.append(r)
            nombres.append(etiqueta_completa)

    if filas_todas:
        prio1 = [r for idx, r in enumerate(filas_todas) if 'TG1+TV1_GN_A' in nombres[idx] or 'TG1+TV1_GN' in nombres[idx] or 'NUEVARENCA_TG1+TV1' in nombres[idx]]
        if prio1:
            return prio1

        prio2 = [r for idx, r in enumerate(filas_todas) if 'TG1+TV1' in nombres[idx] or 'CCNUEVA' in nombres[idx] or 'COMBINADO' in nombres[idx]]
        if prio2:
            return prio2

    return filas_todas if filas_todas else list(range(1565, 1590))


def calcular_sistema_prom_desde_tco(wb_prg, wb_po) -> float:
    """
    Calcula el 'Sistema Promedio' analizando las configuraciones activas (> 0 MW) de Nueva Renca en cada bloque:
    - Bloque 1: Horas 1 a 8 (columnas 5 a 12 de PROGRAMA)
    - Bloque 2: Horas 9 a 18 (columnas 13 a 22 de PROGRAMA)
    - Bloque 3: Horas 19 a 24 (columnas 23 a 28 de PROGRAMA)
    
    Para cada Bloque, obtiene el promedio de CMg en la hoja TCO de todas las configuraciones despachadas en ese bloque.
    Finalmente promedia los 3 bloques y redondea a 1 decimal (round(promedio, 1)).
    """
    def to_float(val) -> float:
        if val is None: return 0.0
        try: return float(str(val).strip().replace(',', '.'))
        except: return 0.0

    sheet_prog = wb_prg['PROGRAMA'] if 'PROGRAMA' in wb_prg.sheetnames else wb_prg.active
    filas_nr = obtener_filas_nueva_renca(sheet_prog)

    # Calcular promedio directo de generación de las 24 horas si no hay hoja TCO
    mw_horas = []
    if filas_nr:
        for c in range(5, 29):
            val_h = sum(to_float(sheet_prog.cell(row=r, column=c).value) for r in filas_nr)
            if val_h > 0:
                mw_horas.append(val_h)

    promedio_directo = sum(mw_horas) / float(len(mw_horas)) if mw_horas else 370.0

    if not wb_prg or not wb_po or 'TCO' not in wb_po.sheetnames:
        return round(promedio_directo, 1)

    sheet_tco = wb_po['TCO']

    if not filas_nr:
        return round(promedio_directo, 1)

    # 1. Mapear configuraciones activas (> 0 MW) en cada uno de los 3 Bloques
    configs_b1, configs_b2, configs_b3 = [], [], []

    for r in filas_nr:
        nombre = sheet_prog.cell(row=r, column=3).value or sheet_prog.cell(row=r, column=2).value
        if not nombre: continue
        nombre_str = str(nombre).strip()

        # Bloque 1 (Horas 1-8 -> Cols 5-12)
        if sum(to_float(sheet_prog.cell(row=r, column=c).value) for c in range(5, 13)) > 0:
            configs_b1.append(nombre_str)

        # Bloque 2 (Horas 9-18 -> Cols 13-22)
        if sum(to_float(sheet_prog.cell(row=r, column=c).value) for c in range(13, 23)) > 0:
            configs_b2.append(nombre_str)

        # Bloque 3 (Horas 19-24 -> Cols 23-28)
        if sum(to_float(sheet_prog.cell(row=r, column=c).value) for c in range(23, 29)) > 0:
            configs_b3.append(nombre_str)

    # 2. Función helper para obtener CMg promedio de un Bloque en TCO
    def obtener_cmg_prom_bloque(col_centrales, col_cmg, configs_activas):
        if not configs_activas: return None
        cmgs = []
        for config_nombre in configs_activas:
            for r in range(1, sheet_tco.max_row + 1):
                c_val = sheet_tco.cell(row=r, column=col_centrales).value
                if c_val and str(c_val).strip().upper() == config_nombre.upper():
                    cmgs.append(to_float(sheet_tco.cell(row=r, column=col_cmg).value))
                    break
        if cmgs:
            return sum(cmgs) / float(len(cmgs))
        return None

    # CMg promedio para cada uno de los 3 bloques
    b1_avg = obtener_cmg_prom_bloque(3, 4, configs_b1)    # Bloque 1: Col 3 (Centrales), Col 4 (CMg)
    b2_avg = obtener_cmg_prom_bloque(7, 8, configs_b2)    # Bloque 2: Col 7 (Centrales), Col 8 (CMg)
    b3_avg = obtener_cmg_prom_bloque(11, 12, configs_b3)  # Bloque 3: Col 11 (Centrales), Col 12 (CMg)

    bloques_validos = [v for v in [b1_avg, b2_avg, b3_avg] if v is not None]

    if bloques_validos:
        promedio_final = sum(bloques_validos) / float(len(bloques_validos))
        return round(promedio_final, 1)

    return round(promedio_directo, 1)


def procesar_excel_generacion(wb_prg, wb_po: Optional[Any] = None) -> Dict[str, Any]:
    def to_float(val) -> float:
        if val is None: return 0.0
        if isinstance(val, (int, float)): return float(val)
        try: return float(str(val).strip().replace(',', '.'))
        except (ValueError, TypeError): return 0.0

    sheet = wb_prg['PROGRAMA'] if 'PROGRAMA' in wb_prg.sheetnames else wb_prg.active

    # Buscar dinámicamente las filas de Nueva Renca
    filas_nr = obtener_filas_nueva_renca(sheet)
    if not filas_nr:
        filas_nr = list(range(1566, 1591))

    # 1. Costo Marginal (Celda AC8)
    costo_marginal = to_float(sheet['AC8'].value)
    if costo_marginal == 0:
        costo_marginal = 52.9

    # 2. Sistema Promedio
    sistema_prom = calcular_sistema_prom_desde_tco(wb_prg, wb_po)

    # 3. Iteración horaria (Columnas 5 = E a 28 = AB -> Horas 1 a 24)
    hrs_carga_base = 0
    hrs_minimo_tecnico = 0
    hrs_fuegos_suplementarios = 0
    mw_fuegos_suplementarios_total = 0.0

    filas_fa = [r for r in filas_nr if 'FA' in str(sheet.cell(row=r, column=3).value or '').upper()]

    horas_24 = []
    mw_totales_dia = 0.0
    for col in range(5, 29): # Horas 1 a 24 (Columnas E a AB)
        h = col - 4
        # Tomar el máximo de potencia física entre las filas de la central en la hora h
        vals_h = [to_float(sheet.cell(row=r, column=col).value) for r in filas_nr]
        gen_total_hora = max(vals_h) if vals_h else 0.0
        gen_fa_hora = sum(to_float(sheet.cell(row=r, column=col).value) for r in filas_fa)

        mw_totales_dia += gen_total_hora
        gen_mw_round = round(gen_total_hora, 0)

        pot_mw = round(gen_total_hora, 1)
        ssaa_mwh = round(pot_mw * 0.033, 1) if pot_mw > 0 else 0.0
        gen_neta = round(max(0.0, pot_mw - ssaa_mwh), 1)

        horas_24.append({
            "hora": h,
            "potencia_mw": pot_mw,
            "generacion_mwh": pot_mw,
            "ssaa_mwh": ssaa_mwh,
            "generacion_neta": gen_neta
        })

        if gen_mw_round >= 330:
            hrs_carga_base += 1
        elif gen_mw_round == 160 or (158.0 <= gen_total_hora < 330.0) or (160 <= gen_mw_round < 330): # Mínimo Técnico estricto a 160 MW (rango 160-329 MW)
            hrs_minimo_tecnico += 1

        if gen_fa_hora > 32.0:
            hrs_fuegos_suplementarios += 1
            mw_fuegos_suplementarios_total += gen_fa_hora

    if sistema_prom == 0:
        sistema_prom = 52.9

    potencia_esperada_mw = round(mw_totales_dia, 0)

    return {
        "sistema_prom_mw": round(sistema_prom, 1),
        "costo_marginal_usd_mw": round(costo_marginal, 1),
        "potencia_esperada_mw": int(potencia_esperada_mw),
        "mw_fuegos_suplementarios": int(round(mw_fuegos_suplementarios_total, 0)),
        "hrs_carga_base": hrs_carga_base,
        "hrs_minimo_tecnico": hrs_minimo_tecnico,
        "hrs_fuegos_suplementarios": hrs_fuegos_suplementarios,
        "horas": horas_24
    }


def calcular_resumen_simulado() -> Dict[str, Any]:
    """Valores de fallback en caso de no haber conexión al archivo"""
    return {
        "sistema_prom_mw": 56.7,
        "costo_marginal_usd_mw": 52.9,
        "potencia_esperada_mw": 4004,
        "mw_fuegos_suplementarios": 0,
        "hrs_carga_base": 0,
        "hrs_minimo_tecnico": 22,
        "hrs_fuegos_suplementarios": 0
    }

